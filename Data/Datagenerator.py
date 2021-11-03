import openpyxl

li=[]
li1=[]
def datagenrator():
    wk=openpyxl.load_workbook("C:\\Users\\SOURAV\\Downloads\\seleniumdata.xlsx")
    sh=wk["Sheet1"]
    r=sh.max_row
    for i in range(1,r+1):
        li1 = []
        un=sh.cell(i,1)
        pd=sh.cell(i,2)
        li1.insert(0,un.value)
        li1.insert(1,pd.value)
        li.insert(i-1,li1)
    print(li1)
    # li=[['us1','us1@gmail.com'],['us2','us2@gmail.com'],['us3','us3@gmail.com']]
    return li